#!/usr/bin/env python3
"""재무제표 3표 → 다운로드용 XLSX 워크북.

data/app/financials.json(= financials.py --all --save 산출)을 읽어
보유 14종목(VOO 제외)의 손익계산서·재무상태표·현금흐름표를 시트별로 편다.

  요약        : 전종목 한 장 비교(매출·마진·FCF·순현금·플래그·펀더 서브스코어)
  플래그       : 재무 경보 상세(측정·경보 전용 — 자동 매매 아님)
  <종목> 시트  : 연간(최대 5기) + 분기(최대 8기) 3표 + 비율

원 수치 = SEC EDGAR XBRL / Yahoo / DART 1차값을 그대로 적재(1차 출처가 정본).
비율·증감률은 그 원수치로부터 계산한 값이며, 산식을 행 이름에 병기해 검산 가능하게 둔다.
(셀 수식이 아니라 값으로 넣는다 — 어떤 뷰어·미리보기에서도 숫자가 그대로 보이게 하기 위함.)

사용: python3 export_financials_xlsx.py [--out <경로>]
"""
from __future__ import annotations

import argparse
import json
import os
from datetime import datetime

# ⚠️ 이 레포에서 유일하게 stdlib이 아닌 스크립트(openpyxl). 컨테이너는 세션마다 새로 뜨므로
#   openpyxl이 없는 게 정상 상태다 → 모듈 최상단에서 import하면 selfcheck 임포트 스모크가
#   매 새 세션마다 GATE FAIL로 죽는다(8/2 실측). 지연 임포트로 바꿔 '스크립트 결함'과
#   '의존성 미설치'를 분리한다 — 실제 실행 시에만 필요하고, 없으면 설치 안내를 내고 종료.
try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    _OPENPYXL_ERR = None
except ModuleNotFoundError as _e:                     # 의존성 미설치 = 정상 상태(새 컨테이너)
    _OPENPYXL_ERR = _e

    class _Missing:                                   # import만 통과시키는 무해한 자리표시자
        def __init__(self, *a, **k): pass
        def __call__(self, *a, **k): return self

    Workbook = Alignment = Border = Font = PatternFill = Side = _Missing

    def get_column_letter(i): return ""


def require_openpyxl():
    """실행 시점에만 의존성을 요구 — '스크립트 결함'과 '의존성 미설치'를 분리한다."""
    if _OPENPYXL_ERR is not None:
        raise SystemExit(
            "❌ openpyxl 미설치 — XLSX 산출물 전용 의존성입니다.\n"
            "   pip install openpyxl  후 다시 실행하세요.\n"
            "   (나머지 파이프라인은 stdlib만 쓰므로 이 스크립트만 영향받습니다.)")

_HERE = os.path.dirname(os.path.abspath(__file__))          # .claude/skills/portfolio-desk/scripts
ROOT = os.path.abspath(os.path.join(_HERE, "..", "..", "..", ".."))  # 프로젝트 루트
SRC = os.path.join(ROOT, "data", "app", "financials.json")

FONT = "Arial"
C_HDR = "1F3864"      # 진남색 헤더
C_BAND = "D9E2F3"     # 블록 구분 밴드
C_INPUT = "0000FF"    # 하드코딩 원수치(파랑)
C_FORMULA = "000000"  # 계산값(검정)
C_WARN = "C00000"

THIN = Side(style="thin", color="BFBFBF")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# (표시명, json 키, 종류)  종류: raw=원수치 / pct=비율 / x=배수
IS_ROWS = [
    ("매출액", "revenue", "raw"),
    ("매출총이익", "gross_profit", "raw"),
    ("영업이익", "operating_income", "raw"),
    ("순이익", "net_income", "raw"),
    ("R&D", "rnd", "raw"),
    ("희석 EPS", "eps_diluted", "eps"),
    ("희석주식수", "shares_diluted", "shares"),
]
BS_ROWS = [
    ("총자산", "assets", "raw"),
    ("총부채", "liabilities", "raw"),
    ("자본총계", "equity", "raw"),
    ("현금성자산", "cash", "raw"),
    ("재고자산", "inventory", "raw"),
    ("매출채권", "receivables", "raw"),
    ("유동자산", "assets_current", "raw"),
    ("유동부채", "liabilities_current", "raw"),
    ("총차입금", "total_debt", "raw"),
    ("계약부채(이연수익)", "deferred_revenue", "raw"),
]
CF_ROWS = [
    ("영업활동 현금흐름(CFO)", "cfo", "raw"),
    ("설비투자(CapEx)", "capex", "raw"),
    ("잉여현금흐름(FCF)", "fcf", "raw"),
    ("자사주매입", "buyback", "raw"),
]
# 비율 = 원수치로부터 계산 (분자키, 분모키)
RATIO_ROWS = [
    ("매출총이익률  (= 매출총이익 ÷ 매출액)", "gross_profit", "revenue", "pct"),
    ("영업이익률  (= 영업이익 ÷ 매출액)", "operating_income", "revenue", "pct"),
    ("순이익률  (= 순이익 ÷ 매출액)", "net_income", "revenue", "pct"),
    ("FCF 마진  (= FCF ÷ 매출액)", "fcf", "revenue", "pct"),
    ("ROE  (= 순이익 ÷ 자본총계)", "net_income", "equity", "pct"),
    ("부채비율  (= 총부채 ÷ 자본총계)", "liabilities", "equity", "x"),
    ("유동비율  (= 유동자산 ÷ 유동부채)", "assets_current", "liabilities_current", "x"),
]

FLAG_KO = {
    "margin_trend_break": "영업마진 추세 하락 전환 — 룰4 메모리 정점 판정 입력",
    "inventory_surge": "재고 증가율이 매출 증가율 초과",
    "fcf_negative_turn": "FCF 마이너스 전환",
    "debt_buildup": "차입 급증 / 순현금 악화",
    "receivable_divergence": "매출채권 증가율이 매출 증가율과 괴리",
    "backlog_growth": "계약부채(수주잔고) 급증 — 강세 플래그",
    "dilution": "주식수 증가(희석)",
    "source_conflict": "두 소스가 1% 초과 불일치 — 병기 확인 필요",
}


def money_fmt(region: str) -> str:
    return '#,##0;(#,##0);-'


def unit_label(region: str) -> str:
    return "백만 USD" if region == "US" else "억원"


def unit_div(region: str) -> float:
    return 1e6 if region == "US" else 1e8


def style_header(ws, row: int, ncol: int, text_first: str, labels: list[str]):
    cells = [text_first] + labels
    for j, v in enumerate(cells, start=1):
        c = ws.cell(row=row, column=j, value=v)
        c.font = Font(name=FONT, bold=True, color="FFFFFF", size=10)
        c.fill = PatternFill("solid", fgColor=C_HDR)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BOX


def band(ws, row: int, ncol: int, title: str):
    for j in range(1, ncol + 1):
        c = ws.cell(row=row, column=j)
        c.fill = PatternFill("solid", fgColor=C_BAND)
        c.border = BOX
    c = ws.cell(row=row, column=1, value=title)
    c.font = Font(name=FONT, bold=True, size=10, color=C_HDR)


def write_company_sheet(wb: Workbook, ticker: str, co: dict):
    name = co.get("name") or ticker
    region = co.get("region", "US")
    cur = co.get("currency", "USD")
    div = unit_div(region)

    annual = co.get("annual", []) or []
    quarterly = co.get("quarterly", []) or []
    # 오래된 → 최신 순으로 (재무제표 관행)
    annual = sorted(annual, key=lambda p: p.get("end", ""))
    quarterly = sorted(quarterly, key=lambda p: p.get("end", ""))
    periods = [("연간", p) for p in annual] + [("분기", p) for p in quarterly]
    if not periods:
        return

    # 국내는 종목코드 대신 한글명이 알아보기 쉽다
    title = (name if region == "KR" else ticker)
    for ch in "[]:*?/\\":
        title = title.replace(ch, "")
    ws = wb.create_sheet(title=title[:28])
    ncol = 1 + len(periods)

    # 타이틀 블록
    t = ws.cell(row=1, column=1, value=f"{name}  ({ticker})")
    t.font = Font(name=FONT, bold=True, size=14, color=C_HDR)
    s = ws.cell(row=2, column=1,
                value=f"단위: {unit_label(region)} · 통화 {cur} · 출처 {co.get('source','')}"
                      f" / 교차 {co.get('cross_source','')} · 기준일 {co.get('as_of','')}")
    s.font = Font(name=FONT, size=9, italic=True, color="595959")
    n2 = ws.cell(row=3, column=1,
                 value="파란 숫자 = 원 재무제표 수치(1차 출처 그대로) · 검정 = 위 원수치로 계산한 값(산식은 행 이름에 병기)")
    n2.font = Font(name=FONT, size=9, italic=True, color="595959")

    hdr_row = 5
    labels = [f"{kind}\n{p.get('end','')}" for kind, p in periods]
    style_header(ws, hdr_row, ncol, "계정과목", labels)
    ws.freeze_panes = ws.cell(row=hdr_row + 1, column=2)

    r = hdr_row + 1
    rowmap: dict[str, int] = {}

    def emit_block(block_title: str, rows):
        nonlocal r
        band(ws, r, ncol, block_title)
        r += 1
        for label, key, kind in rows:
            c = ws.cell(row=r, column=1, value=label)
            c.font = Font(name=FONT, size=10, bold=(kind == "raw" and key in ("revenue", "fcf")))
            c.border = BOX
            for j, (_, p) in enumerate(periods, start=2):
                v = p.get(key)
                cell = ws.cell(row=r, column=j)
                if v is None:
                    cell.value = None
                elif kind == "raw":
                    cell.value = v / div
                    cell.number_format = money_fmt(region)
                elif kind == "eps":
                    cell.value = v
                    cell.number_format = '#,##0.00;(#,##0.00);-'
                elif kind == "shares":
                    cell.value = v / 1e6
                    cell.number_format = '#,##0;(#,##0);-'
                cell.font = Font(name=FONT, size=10, color=C_INPUT)
                cell.border = BOX
            rowmap[key] = r
            r += 1

    emit_block("Ⅰ. 손익계산서", IS_ROWS)
    ws.cell(row=rowmap["shares_diluted"], column=1,
            value="희석주식수 (백만주)").font = Font(name=FONT, size=10)
    r += 1
    emit_block("Ⅱ. 재무상태표", BS_ROWS)
    r += 1
    emit_block("Ⅲ. 현금흐름표", CF_ROWS)
    r += 1

    band(ws, r, ncol, "Ⅳ. 주요 비율 · 파생 (위 원수치로부터 계산)")
    r += 1
    c = ws.cell(row=r, column=1, value="순현금  (= 현금성자산 − 총차입금)")
    c.font = Font(name=FONT, size=10)
    c.border = BOX
    for j, (_, p) in enumerate(periods, start=2):
        cash, debt = p.get("cash"), p.get("total_debt")
        cell = ws.cell(row=r, column=j)
        if cash is not None and debt is not None:
            cell.value = (cash - debt) / div
        cell.number_format = money_fmt(region)
        cell.font = Font(name=FONT, size=10, color=C_FORMULA)
        cell.border = BOX
    r += 1

    for label, num, den, kind in RATIO_ROWS:
        if num not in rowmap or den not in rowmap:
            continue
        c = ws.cell(row=r, column=1, value=label)
        c.font = Font(name=FONT, size=10)
        c.border = BOX
        for j, (_, p) in enumerate(periods, start=2):
            a, b = p.get(num), p.get(den)
            cell = ws.cell(row=r, column=j)
            if a is not None and b:
                cell.value = a / b
            cell.number_format = '0.0%' if kind == "pct" else '0.00"x"'
            cell.font = Font(name=FONT, size=10, color=C_FORMULA)
            cell.border = BOX
        r += 1

    # 매출 YoY (연간 구간만 — 분기는 전분기 대비가 아니라 비교 무의미하므로 생략)
    c = ws.cell(row=r, column=1, value="매출 전기대비 증감률")
    c.font = Font(name=FONT, size=10)
    c.border = BOX
    for j in range(2, ncol + 1):
        cell = ws.cell(row=r, column=j)
        same_kind = j > 2 and periods[j - 2][0] == periods[j - 3][0]
        if same_kind:
            cur, prv = periods[j - 2][1].get("revenue"), periods[j - 3][1].get("revenue")
            if cur is not None and prv:
                cell.value = cur / prv - 1
        cell.number_format = '0.0%'
        cell.font = Font(name=FONT, size=10, color=C_FORMULA)
        cell.border = BOX
    r += 2

    # 플래그 + 서브스코어
    flags = co.get("flags", []) or []
    codes = []
    for f in flags:
        codes.append(f.get("code") if isinstance(f, dict) else f)
    codes = [c for c in codes if c]
    sub = co.get("fund_subscore")
    ws.cell(row=r, column=1, value="재무 플래그").font = Font(name=FONT, bold=True, size=10)
    ws.cell(row=r, column=2,
            value=(", ".join(sorted(set(codes))) if codes else "없음")).font = Font(
        name=FONT, size=10, color=(C_WARN if codes else "000000"))
    r += 1
    ws.cell(row=r, column=1, value="펀더 서브스코어(0~100)").font = Font(name=FONT, bold=True, size=10)
    ws.cell(row=r, column=2, value=sub).font = Font(name=FONT, size=10)
    r += 1
    ws.cell(row=r, column=1,
            value="※ 플래그는 측정·경보 전용이며 자동 매매 신호가 아님.").font = Font(
        name=FONT, size=9, italic=True, color="595959")

    ws.column_dimensions["A"].width = 26
    for j in range(2, ncol + 1):
        ws.column_dimensions[get_column_letter(j)].width = 14
    ws.sheet_view.showGridLines = False


def write_summary(wb: Workbook, data: dict, order: list[str]):
    ws = wb.create_sheet(title="요약", index=0)
    stocks = data["stocks"]

    t = ws.cell(row=1, column=1, value="보유 14종목 재무제표 요약 (VOO=ETF 제외)")
    t.font = Font(name=FONT, bold=True, size=14, color=C_HDR)
    ws.cell(row=2, column=1,
            value=f"생성 {datetime.now().strftime('%Y-%m-%d %H:%M')} · 데이터 기준일 {data.get('updated','')}"
                  f" · 출처 US=SEC EDGAR XBRL / KR=Yahoo·DART").font = Font(
        name=FONT, size=9, italic=True, color="595959")
    ws.cell(row=3, column=1,
            value="금액 단위: 미국=백만 USD · 국내=억원 (최근 연간 기준)").font = Font(
        name=FONT, size=9, italic=True, color="595959")

    cols = ["티커", "종목명", "지역", "결산일", "매출", "영업이익", "순이익",
            "매출총이익률", "영업이익률", "FCF", "순현금", "총차입금",
            "펀더 서브스코어", "재무 플래그"]
    hdr = 5
    for j, v in enumerate(cols, start=1):
        c = ws.cell(row=hdr, column=j, value=v)
        c.font = Font(name=FONT, bold=True, color="FFFFFF", size=10)
        c.fill = PatternFill("solid", fgColor=C_HDR)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BOX

    r = hdr + 1
    for tk in order:
        co = stocks.get(tk)
        if not co:
            continue
        annual = sorted(co.get("annual", []) or [], key=lambda p: p.get("end", ""))
        if not annual:
            continue
        p = annual[-1]
        region = co.get("region", "US")
        div = unit_div(region)
        codes = sorted({(f.get("code") if isinstance(f, dict) else f) for f in (co.get("flags") or []) if f})

        vals = [
            tk, co.get("name", ""), region, p.get("end", ""),
            (p.get("revenue") or 0) / div,
            (p.get("operating_income") or 0) / div,
            (p.get("net_income") or 0) / div,
            None, None,
            (p.get("fcf") or 0) / div,
            None,
            (p.get("total_debt") or 0) / div,
            co.get("fund_subscore"),
            ", ".join(codes) if codes else "없음",
        ]
        for j, v in enumerate(vals, start=1):
            c = ws.cell(row=r, column=j, value=v)
            c.font = Font(name=FONT, size=10,
                          color=(C_WARN if (j == 14 and codes) else C_INPUT if j in (5, 6, 7, 10, 12) else "000000"))
            c.border = BOX
            if j in (5, 6, 7, 10, 12):
                c.number_format = money_fmt(region)
        # 비율·순현금 = 수식
        c = ws.cell(row=r, column=8, value=p.get("gross_margin"))
        c.number_format = '0.0%'
        c.font = Font(name=FONT, size=10)
        c.border = BOX
        rev, oi = p.get("revenue"), p.get("operating_income")
        c = ws.cell(row=r, column=9, value=(oi / rev) if (oi is not None and rev) else None)
        c.number_format = '0.0%'
        c.font = Font(name=FONT, size=10)
        c.border = BOX
        nc = p.get("net_cash")
        c = ws.cell(row=r, column=11, value=(nc / div) if nc is not None else None)
        c.number_format = money_fmt(region)
        c.font = Font(name=FONT, size=10, color=C_INPUT)
        c.border = BOX
        r += 1

    ws.cell(row=r + 1, column=1,
            value="※ 미국(백만 USD)과 국내(억원)는 단위가 달라 세로 합산 불가 — 종목 간 비교는 비율 컬럼으로 할 것.").font = Font(
        name=FONT, size=9, italic=True, color=C_WARN)
    ws.cell(row=r + 2, column=1,
            value="※ 재무 플래그는 측정·경보 전용. 자동 매매 신호가 아니며 최종 판단은 정훈.").font = Font(
        name=FONT, size=9, italic=True, color="595959")

    widths = [12, 22, 7, 12, 14, 14, 14, 12, 12, 14, 14, 14, 12, 34]
    for j, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(j)].width = w
    ws.freeze_panes = ws.cell(row=hdr + 1, column=3)
    ws.sheet_view.showGridLines = False


def write_flags(wb: Workbook, data: dict, order: list[str]):
    ws = wb.create_sheet(title="플래그")
    ws.cell(row=1, column=1, value="재무 플래그 상세 — 측정·경보 전용(자동 매매 아님)").font = Font(
        name=FONT, bold=True, size=14, color=C_HDR)
    cols = ["티커", "종목명", "플래그", "의미", "상세"]
    hdr = 3
    for j, v in enumerate(cols, start=1):
        c = ws.cell(row=hdr, column=j, value=v)
        c.font = Font(name=FONT, bold=True, color="FFFFFF", size=10)
        c.fill = PatternFill("solid", fgColor=C_HDR)
        c.border = BOX
        c.alignment = Alignment(horizontal="center", wrap_text=True)
    r = hdr + 1
    for tk in order:
        co = data["stocks"].get(tk)
        if not co:
            continue
        for f in (co.get("flags") or []):
            code = f.get("code") if isinstance(f, dict) else f
            detail = (f.get("message") or f.get("detail") or "") if isinstance(f, dict) else ""
            for j, v in enumerate([tk, co.get("name", ""), code, FLAG_KO.get(code, ""), detail], start=1):
                c = ws.cell(row=r, column=j, value=v)
                c.font = Font(name=FONT, size=10, color=(C_WARN if j == 3 else "000000"))
                c.border = BOX
                c.alignment = Alignment(wrap_text=(j >= 4), vertical="top")
            r += 1
    for j, w in enumerate([12, 22, 24, 44, 60], start=1):
        ws.column_dimensions[get_column_letter(j)].width = w
    ws.sheet_view.showGridLines = False


def write_subscore(wb: Workbook, data: dict, order: list[str]):
    """펀더 서브스코어 5축 분해 — CANSLIM의 C·A + 마진방향·FCF·재무건전성."""
    ws = wb.create_sheet(title="펀더스코어")
    ws.cell(row=1, column=1, value="펀더 서브스코어 0~100 — CANSLIM 5축 자동 채점").font = Font(
        name=FONT, bold=True, size=14, color=C_HDR)
    ws.cell(row=2, column=1,
            value="N(신제품·촉매)·L(주도주)·M(시장방향)은 자동 채점 대상이 아니며 PM 판단 몫. "
                  "보고서 0~100 스코어와 25점 이상 벌어지면 그 이격을 근거로 설명해야 한다.").font = Font(
        name=FONT, size=9, italic=True, color="595959")

    axes = ["C_분기EPS", "A_연간EPS", "마진방향", "FCF건전성", "재무건전성"]
    hdr = 4
    cols = ["티커", "종목명", "합계"] + [a for a in axes] + ["채점 근거(마진방향)"]
    for j, v in enumerate(cols, start=1):
        c = ws.cell(row=hdr, column=j, value=v)
        c.font = Font(name=FONT, bold=True, color="FFFFFF", size=10)
        c.fill = PatternFill("solid", fgColor=C_HDR)
        c.alignment = Alignment(horizontal="center", wrap_text=True)
        c.border = BOX

    r = hdr + 1
    for tk in order:
        co = data["stocks"].get(tk)
        if not co:
            continue
        ax = co.get("axes") or {}
        ws.cell(row=r, column=1, value=tk).border = BOX
        ws.cell(row=r, column=2, value=co.get("name", "")).border = BOX
        # 합계 = 5축 수식 합
        total = sum((ax.get(a) or {}).get("score") or 0 for a in axes)
        c = ws.cell(row=r, column=3, value=round(total, 1))
        c.font = Font(name=FONT, bold=True, size=10)
        c.number_format = '0.0'
        c.border = BOX
        for j, a in enumerate(axes, start=4):
            v = (ax.get(a) or {}).get("score")
            cell = ws.cell(row=r, column=j, value=v)
            cell.number_format = '0.0'
            cell.font = Font(name=FONT, size=10, color=C_INPUT)
            cell.border = BOX
        why = (ax.get("마진방향") or {}).get("why", "")
        c = ws.cell(row=r, column=9, value=why)
        c.font = Font(name=FONT, size=9)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        c.border = BOX
        for j in range(1, 3):
            ws.cell(row=r, column=j).font = Font(name=FONT, size=10)
        r += 1

    ws.cell(row=r + 1, column=1,
            value="배점: C 분기EPS 25 · A 연간EPS 25 · 마진방향 20 · FCF건전성 15 · 재무건전성 15").font = Font(
        name=FONT, size=9, italic=True, color="595959")
    for j, w in enumerate([12, 22, 10, 12, 12, 12, 12, 12, 46], start=1):
        ws.column_dimensions[get_column_letter(j)].width = w
    ws.freeze_panes = ws.cell(row=hdr + 1, column=3)
    ws.sheet_view.showGridLines = False


def main():
    require_openpyxl()
    ap = argparse.ArgumentParser()
    ap.add_argument("--out", default=os.path.join(ROOT, "output", "재무제표_보유14종목.xlsx"))
    ap.add_argument("--src", default=SRC)
    args = ap.parse_args()

    with open(args.src, encoding="utf-8") as fh:
        data = json.load(fh)

    order = ["005930.KS", "066570.KS", "454910.KS", "005380.KS", "035420.KS",
             "NVDA", "MU", "AAPL", "MSFT", "ANET", "AVGO", "GOOGL", "META", "ORCL"]
    order = [t for t in order if t in data["stocks"]] + \
            [t for t in data["stocks"] if t not in order]

    wb = Workbook()
    wb.remove(wb.active)
    write_summary(wb, data, order)
    write_subscore(wb, data, order)
    write_flags(wb, data, order)
    for tk in order:
        write_company_sheet(wb, tk, data["stocks"][tk])

    os.makedirs(os.path.dirname(args.out), exist_ok=True)
    wb.save(args.out)
    print(f"저장: {args.out}")
    print(f"시트 {len(wb.sheetnames)}개: {', '.join(wb.sheetnames)}")


if __name__ == "__main__":
    main()
